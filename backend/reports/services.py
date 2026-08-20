import csv
import io
import calendar
from django.http import HttpResponse
from income.repositories import IncomeRepository
from expenses.repositories import ExpenseRepository


class ReportService:
    """
    Service for generating downloadable reports (CSV, Excel).
    """

    @staticmethod
    def _get_transactions(user, year=None, month=None):
        filters = {}
        if year:
            if month:
                last_day = calendar.monthrange(year, month)[1]
                filters['date_from'] = f"{year}-{month:02d}-01"
                filters['date_to'] = f"{year}-{month:02d}-{last_day:02d}"
            else:
                filters['date_from'] = f"{year}-01-01"
                filters['date_to'] = f"{year}-12-31"

        incomes = IncomeRepository.get_all_for_user(user, filters)
        expenses = ExpenseRepository.get_all_for_user(user, filters)

        transactions = []
        for inc in incomes:
            cat_name = inc.category.name if inc.category else "None"
            transactions.append({
                'date': inc.date,
                'type': 'Income',
                'category': cat_name,
                'title': inc.title,
                'amount': float(inc.amount),
                'desc': inc.description
            })

        for exp in expenses:
            cat_name = exp.category.name if exp.category else "None"
            transactions.append({
                'date': exp.date,
                'type': 'Expense',
                'category': cat_name,
                'title': f"{exp.title} ({exp.payee})" if exp.payee else exp.title,
                'amount': -float(exp.amount),
                'desc': exp.description
            })

        transactions.sort(key=lambda x: x['date'])
        return transactions

    @staticmethod
    def generate_csv_report(user, year=None, month=None):
        transactions = ReportService._get_transactions(user, year, month)

        response = HttpResponse(content_type='text/csv')
        filename = f"financial_report_{year or 'all'}_{month or 'all'}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(['Date', 'Type', 'Category', 'Title / Payee', 'Amount (USD)', 'Description'])

        for t in transactions:
            writer.writerow([
                t['date'].strftime('%Y-%m-%d'),
                t['type'],
                t['category'],
                t['title'],
                t['amount'],
                t['desc']
            ])

        return response

    @staticmethod
    def generate_excel_report(user, year=None, month=None):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        transactions = ReportService._get_transactions(user, year, month)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Report"

        # Header styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        headers = ['Date', 'Type', 'Category', 'Title / Payee', 'Amount (USD)', 'Description']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Data rows
        income_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        expense_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        for row_num, t in enumerate(transactions, 2):
            row_fill = income_fill if t['type'] == 'Income' else expense_fill
            row_data = [
                t['date'].strftime('%Y-%m-%d'),
                t['type'],
                t['category'],
                t['title'],
                t['amount'],
                t['desc']
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = row_fill
                cell.border = thin_border

        # Auto-fit columns
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_length + 4, 50)

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        total_income = sum(t['amount'] for t in transactions if t['type'] == 'Income')
        total_expense = abs(sum(t['amount'] for t in transactions if t['type'] == 'Expense'))
        net = total_income - total_expense

        summary_data = [
            ("Total Income", total_income),
            ("Total Expenses", total_expense),
            ("Net Savings", net),
        ]
        ws2.cell(row=1, column=1, value="Metric").font = header_font
        ws2.cell(row=1, column=1).fill = header_fill
        ws2.cell(row=1, column=2, value="Amount (USD)").font = header_font
        ws2.cell(row=1, column=2).fill = header_fill

        for i, (label, val) in enumerate(summary_data, 2):
            ws2.cell(row=i, column=1, value=label)
            ws2.cell(row=i, column=2, value=round(val, 2))

        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 18

        # Write to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"financial_report_{year or 'all'}_{month or 'all'}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
